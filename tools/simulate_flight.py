# -------------------------------------------------------------
# # Project : FlightSoftware-Mastodonte
# # File    : simulate_flight.py
# # Author  : Paul Miailhe / Antigravity
# # Date    : 2026-06-17
# # Object  : Simulation du vol et du Filtre de Kalman 1D
# -------------------------------------------------------------

import os
import shutil
import openpyxl
import matplotlib.pyplot as plt

# =====================================================================
# AJUSTEMENT DE L'INDICE DE CONFIANCE DU FILTRE DE KALMAN
# =====================================================================
# Vous pouvez ajuster les paramètres ci-dessous pour modifier la réponse du filtre :
#
# - Q_alt / Q_vel (Bruit de Process) : Confiance dans le modèle physique de Newton.
#   Plus ces valeurs sont grandes, plus le filtre accepte que le modèle dévie,
#   et donc plus il va faire confiance aux mesures réelles du capteur (réactif mais moins filtré).
#
# - R_alt (Bruit de Mesure) : Confiance dans le capteur (baromètre).
#   Plus cette valeur est grande, plus le filtre estime que le capteur est bruité.
#   Le filtre va alors lisser fortement la courbe en se basant sur sa prédiction (très lisse, mais introduit du retard).
#
# Règle d'or : 
#   * Augmenter R_alt / Q_vel  ==> Plus de filtrage (lisse), mais plus de retard de détection.
#   * Diminuer R_alt / Q_vel  ==> Moins de retard (réactif), mais plus sensible aux bruits et fausses apogées.
# =====================================================================

Q_alt = 0.1   # Variance de l'incertitude sur l'altitude théorique
Q_vel = 0.5   # Variance de l'incertitude sur la vitesse théorique
R_alt = 1.0   # Variance du bruit de mesure du baromètre (en mètres carrés)

# =====================================================================

# Fichiers de données (l'excel est situé dans le dossier test/)
excel_path = os.path.join(os.path.dirname(__file__), "..", "test", "stabtraj_v3-5-3.xlsx")
temp_path = os.path.join(os.path.dirname(__file__), "temp_stabtraj.xlsx")

# Vérification de l'existence du fichier
if not os.path.exists(excel_path):
    # Essayons dans le dossier courant si jamais le script a été déplacé différemment
    excel_path = os.path.join(os.path.dirname(__file__), "stabtraj_v3-5-3.xlsx")

if not os.path.exists(excel_path):
    raise FileNotFoundError(f"Impossible de trouver le fichier stabtraj_v3-5-3.xlsx (cherché dans {excel_path})")

# Copie temporaire pour contourner le verrouillage d'écriture d'Excel
if os.path.exists(temp_path):
    try:
        os.remove(temp_path)
    except Exception:
        pass

try:
    shutil.copy2(excel_path, temp_path)
except Exception:
    import subprocess
    # En cas de verrouillage d'écriture de Windows/Excel, PowerShell contourne l'erreur
    subprocess.run(["powershell", "-Command", f"Copy-Item -Path '{excel_path}' -Destination '{temp_path}' -Force"], capture_output=True)


# Chargement du classeur
wb = openpyxl.load_workbook(temp_path, data_only=True)
sheet = wb["Calculs"]

raw_times = []
raw_altitudes = []
raw_velocities_z = []

max_row = sheet.max_row
for r_idx in range(4, max_row + 1):
    t = sheet.cell(row=r_idx, column=2).value
    z = sheet.cell(row=r_idx, column=11).value
    vz = sheet.cell(row=r_idx, column=8).value
    if t is not None and z is not None:
        try:
            raw_times.append(float(t))
            raw_altitudes.append(float(z))
            raw_velocities_z.append(float(vz) if vz is not None else 0.0)
        except ValueError:
            pass

print(f"Chargement de {len(raw_times)} points depuis Excel.")

# Interpolation à 25 Hz (pas de temps constant dt = 40 ms)
dt = 0.040
t_start = 0.0
t_end = raw_times[-1]

times = []
altitudes = []
velocities_z_actual = []

t_curr = t_start
raw_idx = 0

while t_curr <= t_end:
    while raw_idx < len(raw_times) - 1 and raw_times[raw_idx + 1] < t_curr:
        raw_idx += 1
        
    if raw_idx >= len(raw_times) - 1:
        z_interp = raw_altitudes[-1]
        vz_interp = raw_velocities_z[-1]
    else:
        t0 = raw_times[raw_idx]
        t1 = raw_times[raw_idx + 1]
        z0 = raw_altitudes[raw_idx]
        z1 = raw_altitudes[raw_idx + 1]
        vz0 = raw_velocities_z[raw_idx]
        vz1 = raw_velocities_z[raw_idx + 1]
        
        ratio = (t_curr - t0) / (t1 - t0)
        z_interp = z0 + ratio * (z1 - z0)
        vz_interp = vz0 + ratio * (vz1 - vz0)
        
    times.append(t_curr)
    altitudes.append(z_interp)
    velocities_z_actual.append(vz_interp)
    t_curr += dt

t_last = times[-1]

# Ajout de 10 secondes d'immobilité à la fin pour valider l'atterrissage
for i in range(1, 251):
    times.append(t_last + i * dt)
    altitudes.append(0.0)
    velocities_z_actual.append(0.0)

# Initialisation du filtre de Kalman
kalman_z = 0.0
kalman_v = 0.0
P_cov = [[1.0, 0.0], [0.0, 1.0]]

currentState = "ASCEND"
apogee_counter = 0
touchdown_confirm_counter = 0

triggerBaroApogee = False
triggerTouch = False

apogee_time_detected = 0.0
apogee_alt_detected = 0.0
touchdown_time_detected = 0.0

history_t = []
history_z_meas = []
history_z_kal = []
history_v_act = []
history_v_kal = []
history_states = []

for idx in range(len(times)):
    t = times[idx]
    z_meas = altitudes[idx]
    vz_actual = velocities_z_actual[idx]
    
    # 1. Prediction
    z_pred = kalman_z + (kalman_v * dt)
    v_pred = kalman_v

    # 2. Covariance Prediction
    P_cov[0][0] = P_cov[0][0] + dt * (P_cov[1][0] + P_cov[0][1]) + dt * dt * P_cov[1][1] + Q_alt
    P_cov[0][1] = P_cov[0][1] + dt * P_cov[1][1]
    P_cov[1][0] = P_cov[0][1]
    P_cov[1][1] = P_cov[1][1] + Q_vel

    # 3. Gain de Kalman
    S = P_cov[0][0] + R_alt
    K0 = P_cov[0][0] / S
    K1 = P_cov[1][0] / S

    # 4. Innovation
    y = z_meas - z_pred

    # 5. Correction
    kalman_z = z_pred + (K0 * y)
    kalman_v = v_pred + (K1 * y)

    # 6. Covariance Update
    P_cov[0][0] = (1.0 - K0) * P_cov[0][0]
    P_cov[0][1] = (1.0 - K0) * P_cov[0][1]
    P_cov[1][0] = P_cov[0][1]
    P_cov[1][1] = P_cov[1][1] - K1 * P_cov[0][1]
    
    # Détection Apogée
    if currentState == "ASCEND":
        if kalman_z > 15.0 and kalman_v < -1.0:
            apogee_counter += 1
            if apogee_counter >= 5:
                if not triggerBaroApogee:
                    triggerBaroApogee = True
                    apogee_time_detected = t
                    apogee_alt_detected = kalman_z
                    currentState = "DESCEND"
        else:
            apogee_counter = 0

    # Détection Touchdown
    elif currentState == "DESCEND":
        if abs(kalman_v) < 0.5:
            touchdown_confirm_counter += 1
            if touchdown_confirm_counter >= 50:
                if not triggerTouch:
                    triggerTouch = True
                    touchdown_time_detected = t
                    currentState = "TOUCHDOWN"
        else:
            touchdown_confirm_counter = 0
            
    history_t.append(t)
    history_z_meas.append(z_meas)
    history_z_kal.append(kalman_z)
    history_v_act.append(vz_actual)
    history_v_kal.append(kalman_v)
    history_states.append(currentState)

# Nettoyage fichier temporaire
try:
    os.remove(temp_path)
except Exception:
    pass

# =====================================================================
# RAPPORTS ET RESULTATS
# =====================================================================
actual_apogee = max(raw_altitudes)
actual_apogee_time = raw_times[raw_altitudes.index(actual_apogee)]

print("\n" + "="*40)
print("     RESULTATS DE LA SIMULATION KALMAN")
print("="*40)
print(f"Paramètres utilisés : Q_alt={Q_alt}, Q_vel={Q_vel}, R_alt={R_alt}")
print(f"Apogée réelle Excel   : {actual_apogee:.2f} m à t = {actual_apogee_time:.2f} s")

if triggerBaroApogee:
    delay = apogee_time_detected - actual_apogee_time
    err = actual_apogee - apogee_alt_detected
    print(f"Détection de l'apogée : {apogee_alt_detected:.2f} m à t = {apogee_time_detected:.2f} s")
    print(f"  -> Retard de déclenchement : {delay:.3f} s")
    print(f"  -> Perte d'altitude        : {err:.2f} m")
else:
    print("Détection de l'apogée : ECHEC !")

if triggerTouch:
    print(f"Détection de l'atterrissage à t = {touchdown_time_detected:.2f} s (impact à t = {t_last:.2f}s)")
    print(f"  -> Retard après impact : {touchdown_time_detected - t_last:.2f} s")
else:
    print("Détection de l'atterrissage : ECHEC !")

# =====================================================================
# TRACE DES GRAPHES
# =====================================================================
print("\nGénération des courbes de vol...")

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 8), sharex=True)

# 1. Courbe d'Altitude
ax1.plot(history_t, history_z_meas, label="Mesure brute (Excel)", color="gray", linestyle="dashed", alpha=0.7)
ax1.plot(history_t, history_z_kal, label="Filtre de Kalman (Altitude)", color="blue", linewidth=2)
if triggerBaroApogee:
    ax1.plot(apogee_time_detected, apogee_alt_detected, 'ro', markersize=8, label="Déclenchement Apogée")
    ax1.axvline(apogee_time_detected, color="red", linestyle=":", alpha=0.5)
ax1.set_ylabel("Altitude (m)")
ax1.set_title("Estimation de l'Altitude et de la Vitesse via Kalman")
ax1.grid(True)
ax1.legend()

# 2. Courbe de Vitesse
ax2.plot(history_t, history_v_act, label="Vitesse réelle (Excel)", color="gray", linestyle="dashed", alpha=0.7)
ax2.plot(history_t, history_v_kal, label="Filtre de Kalman (Vitesse)", color="green", linewidth=2)
ax2.axhline(0, color="black", linestyle="-", linewidth=0.5)
if triggerBaroApogee:
    ax2.plot(apogee_time_detected, history_v_kal[history_t.index(apogee_time_detected)], 'ro', markersize=8)
    ax2.axvline(apogee_time_detected, color="red", linestyle=":", alpha=0.5)
if triggerTouch:
    ax2.plot(touchdown_time_detected, history_v_kal[history_t.index(touchdown_time_detected)], 'go', markersize=8, label="Touchdown Détecté")
    ax2.axvline(touchdown_time_detected, color="green", linestyle=":", alpha=0.5)
ax2.axhline(-1.0, color="red", linestyle="--", alpha=0.5, label="Seuil Vitesse Apogée (-1 m/s)")
ax2.set_xlabel("Temps (s)")
ax2.set_ylabel("Vitesse (m/s)")
ax2.grid(True)
ax2.legend()

# Sauvegarde de l'image dans le même dossier
output_img = os.path.join(os.path.dirname(__file__), "flight_simulation.png")
plt.tight_layout()
plt.savefig(output_img, dpi=150)
print(f"Courbes enregistrées avec succès : {output_img}")
