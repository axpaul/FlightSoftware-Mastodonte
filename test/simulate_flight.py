import os
import openpyxl
import math

temp_path = r"C:\Users\paulm\.gemini\antigravity\brain\1d5eea88-4085-4f35-9c8b-aae63bf889e7\scratch\temp_stabtraj.xlsx"

wb = openpyxl.load_workbook(temp_path, data_only=True)
sheet = wb["Calculs"]

raw_times = []
raw_altitudes = []
raw_velocities_z = []

max_row = sheet.max_row
for r_idx in range(4, max_row + 1):
    t = sheet.cell(row=r_idx, column=2).value
    z = sheet.cell(row=r_idx, column=11).value
    vz = sheet.cell(row=r_idx, column=8).value
    if t is not None and z is not None:
        try:
            raw_times.append(float(t))
            raw_altitudes.append(float(z))
            raw_velocities_z.append(float(vz) if vz is not None else 0.0)
        except ValueError:
            pass

dt = 0.040
t_start = 0.0
t_end = raw_times[-1]

times = []
altitudes = []
velocities_z_actual = []

t_curr = t_start
raw_idx = 0

while t_curr <= t_end:
    while raw_idx < len(raw_times) - 1 and raw_times[raw_idx + 1] < t_curr:
        raw_idx += 1
        
    if raw_idx >= len(raw_times) - 1:
        z_interp = raw_altitudes[-1]
        vz_interp = raw_velocities_z[-1]
    else:
        t0 = raw_times[raw_idx]
        t1 = raw_times[raw_idx + 1]
        z0 = raw_altitudes[raw_idx]
        z1 = raw_altitudes[raw_idx + 1]
        vz0 = raw_velocities_z[raw_idx]
        vz1 = raw_velocities_z[raw_idx + 1]
        
        ratio = (t_curr - t0) / (t1 - t0)
        z_interp = z0 + ratio * (z1 - z0)
        vz_interp = vz0 + ratio * (vz1 - vz0)
        
    times.append(t_curr)
    altitudes.append(z_interp)
    velocities_z_actual.append(vz_interp)
    t_curr += dt

t_last = times[-1]
# Add 10.0 seconds of ground state to give plenty of time
for i in range(1, 251): # 250 ticks = 10s
    times.append(t_last + i * dt)
    altitudes.append(0.0)
    velocities_z_actual.append(0.0)

# Kalman parameters
kalman_z = 0.0
kalman_v = 0.0
P_cov = [[1.0, 0.0], [0.0, 1.0]]
Q_alt = 0.1
Q_vel = 0.5
R_alt = 1.0

currentState = "ASCEND"
apogee_counter = 0
touchdown_confirm_counter = 0

triggerBaroApogee = False
triggerTouch = False

apogee_time_detected = None
apogee_alt_detected = None
touchdown_time_detected = None

history = []

for idx in range(len(times)):
    t = times[idx]
    z_meas = altitudes[idx]
    vz_actual = velocities_z_actual[idx]
    
    # Kalman update
    z_pred = kalman_z + (kalman_v * dt)
    v_pred = kalman_v

    P_cov[0][0] = P_cov[0][0] + dt * (P_cov[1][0] + P_cov[0][1]) + dt * dt * P_cov[1][1] + Q_alt
    P_cov[0][1] = P_cov[0][1] + dt * P_cov[1][1]
    P_cov[1][0] = P_cov[0][1]
    P_cov[1][1] = P_cov[1][1] + Q_vel

    S = P_cov[0][0] + R_alt
    K0 = P_cov[0][0] / S
    K1 = P_cov[1][0] / S

    y = z_meas - z_pred

    kalman_z = z_pred + (K0 * y)
    kalman_v = v_pred + (K1 * y)

    P_cov[0][0] = (1.0 - K0) * P_cov[0][0]
    P_cov[0][1] = (1.0 - K0) * P_cov[0][1]
    P_cov[1][0] = P_cov[0][1]
    P_cov[1][1] = P_cov[1][1] - K1 * P_cov[0][1]
    
    # Apogee Detection
    if currentState == "ASCEND":
        if kalman_z > 15.0 and kalman_v < -1.0:
            apogee_counter += 1
            if apogee_counter >= 5:
                if not triggerBaroApogee:
                    triggerBaroApogee = True
                    apogee_time_detected = t
                    apogee_alt_detected = kalman_z
                    currentState = "DESCEND"
        else:
            apogee_counter = 0

    # Touchdown Detection
    elif currentState == "DESCEND":
        if abs(kalman_v) < 0.5:
            touchdown_confirm_counter += 1
            if touchdown_confirm_counter >= 50:
                if not triggerTouch:
                    triggerTouch = True
                    touchdown_time_detected = t
                    currentState = "TOUCHDOWN"
        else:
            touchdown_confirm_counter = 0
            
    history.append({
        't': t,
        'z_meas': z_meas,
        'vz_actual': vz_actual,
        'kalman_z': kalman_z,
        'kalman_v': kalman_v,
        'apogee_counter': apogee_counter,
        'touchdown_cnt': touchdown_confirm_counter,
        'state': currentState
    })

# Print values around touchdown (from t_last - 0.5s to t_last + 5.0s)
print("\n=== VALUES AROUND IMPACT & RECOVERY ===")
print(f"{'t (s)':<8} | {'z_meas (m)':<12} | {'kalman_z':<12} | {'kalman_v':<12} | {'touch_cnt':<10} | {'state':<10}")
print("-" * 75)
for h in history:
    if t_last - 0.5 <= h['t'] <= t_last + 6.0:
        print(f"{h['t']:<8.2f} | {h['z_meas']:<12.2f} | {h['kalman_z']:<12.2f} | {h['kalman_v']:<12.2f} | {h['touchdown_cnt']:<10} | {h['state']:<10}")

print("\n=== SIMULATION SUMMARY ===")
if triggerTouch:
    print(f"Detected Touchdown at t = {touchdown_time_detected:.2f} s (impact occurred at t = {t_last:.2f}s)")
    print(f"  -> Touchdown detection delay after impact: {touchdown_time_detected - t_last:.2f} s")
else:
    print("Detected Touchdown: FAILED TO DETECT!")
